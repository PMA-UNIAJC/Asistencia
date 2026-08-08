import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import io

# ============================================================
# VALIDACIÓN DE ARCHIVOS
# ============================================================
def validar_archivo_xlsx(file_obj):
    nombre = getattr(file_obj, "name", "")
    if nombre and not str(nombre).lower().endswith(".xlsx"):
        raise Exception(
            "Formato no soportado. Para esta versión en Streamlit Cloud solo se aceptan archivos .xlsx."
        )

# ============================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================
st.set_page_config(
    page_title="Procesador Académico",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# ESTILOS CSS
# ============================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;600;700&family=JetBrains+Mono:wght@400;600&display=swap');

html, body, [class*="css"] {
    font-family: 'Sora', sans-serif;
}

/* Fondo general */
.stApp {
    background: linear-gradient(135deg, #0f0f1a 0%, #1a1a2e 50%, #16213e 100%);
    min-height: 100vh;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0d1117 0%, #161b27 100%);
    border-right: 1px solid rgba(99, 179, 237, 0.15);
}

[data-testid="stSidebar"] * {
    font-family: 'Sora', sans-serif !important;
}

/* Título principal en sidebar */
.sidebar-title {
    font-family: 'Sora', sans-serif;
    font-size: 1.3rem;
    font-weight: 700;
    color: #63b3ed;
    padding: 1rem 0 0.5rem 0;
    letter-spacing: 0.02em;
}

.sidebar-subtitle {
    font-size: 0.75rem;
    color: rgba(255,255,255,0.4);
    margin-bottom: 1.5rem;
    letter-spacing: 0.05em;
    text-transform: uppercase;
}

/* Radio buttons del sidebar */
[data-testid="stSidebar"] .stRadio label {
    color: rgba(255,255,255,0.75) !important;
    font-size: 0.9rem !important;
    padding: 0.4rem 0 !important;
    cursor: pointer;
    transition: color 0.2s;
}

[data-testid="stSidebar"] .stRadio label:hover {
    color: #63b3ed !important;
}

/* Encabezado de módulo */
.module-header {
    background: linear-gradient(135deg, rgba(99,179,237,0.08) 0%, rgba(154,117,234,0.08) 100%);
    border: 1px solid rgba(99,179,237,0.2);
    border-radius: 16px;
    padding: 2rem 2.5rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}

.module-header::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: linear-gradient(90deg, #63b3ed, #9a75ea, #63b3ed);
}

.module-header h1 {
    font-family: 'Sora', sans-serif;
    font-size: 1.9rem;
    font-weight: 700;
    color: #e2e8f0;
    margin: 0 0 0.4rem 0;
    letter-spacing: -0.01em;
}

.module-header p {
    color: rgba(255,255,255,0.5);
    font-size: 0.9rem;
    margin: 0;
    line-height: 1.6;
}

.module-icon {
    font-size: 2.5rem;
    margin-bottom: 0.8rem;
    display: block;
}

/* Tarjeta de instrucciones */
.info-card {
    background: rgba(99,179,237,0.06);
    border: 1px solid rgba(99,179,237,0.15);
    border-left: 3px solid #63b3ed;
    border-radius: 10px;
    padding: 1rem 1.4rem;
    margin-bottom: 1.5rem;
    color: rgba(255,255,255,0.7);
    font-size: 0.87rem;
    line-height: 1.7;
}

/* Tarjeta de advertencia */
.warning-card {
    background: rgba(246,173,85,0.06);
    border: 1px solid rgba(246,173,85,0.2);
    border-left: 3px solid #f6ad55;
    border-radius: 10px;
    padding: 1rem 1.4rem;
    margin-bottom: 1.5rem;
    color: rgba(255,255,255,0.7);
    font-size: 0.87rem;
    line-height: 1.7;
}

/* Sección de upload */
.upload-section {
    background: rgba(255,255,255,0.02);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 12px;
    padding: 1.5rem;
    margin-bottom: 1.2rem;
}

.upload-label {
    font-size: 0.82rem;
    font-weight: 600;
    color: rgba(255,255,255,0.5);
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 0.5rem;
}

/* File uploader */
[data-testid="stFileUploader"] {
    background: rgba(255,255,255,0.02) !important;
    border: 1.5px dashed rgba(99,179,237,0.25) !important;
    border-radius: 10px !important;
    transition: border-color 0.2s !important;
}

[data-testid="stFileUploader"]:hover {
    border-color: rgba(99,179,237,0.5) !important;
}

/* Botón principal */
.stButton > button {
    background: linear-gradient(135deg, #2b6cb0 0%, #553c9a 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.7rem 2.5rem !important;
    font-family: 'Sora', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    letter-spacing: 0.02em !important;
    transition: all 0.25s ease !important;
    box-shadow: 0 4px 15px rgba(43,108,176,0.3) !important;
}

.stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 25px rgba(43,108,176,0.5) !important;
}

.stButton > button:active {
    transform: translateY(0) !important;
}

/* Download button */
[data-testid="stDownloadButton"] button {
    background: linear-gradient(135deg, #276749 0%, #2f855a 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.7rem 2.5rem !important;
    font-family: 'Sora', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    transition: all 0.25s ease !important;
    box-shadow: 0 4px 15px rgba(39,103,73,0.4) !important;
}

[data-testid="stDownloadButton"] button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 25px rgba(39,103,73,0.6) !important;
}

/* Checkboxes */
.stCheckbox label {
    color: rgba(255,255,255,0.8) !important;
    font-size: 0.9rem !important;
}

/* Mensajes de éxito / error */
.stSuccess {
    background: rgba(39,103,73,0.15) !important;
    border: 1px solid rgba(72,187,120,0.3) !important;
    border-radius: 10px !important;
    color: #9ae6b4 !important;
}

.stError {
    background: rgba(154,44,44,0.15) !important;
    border: 1px solid rgba(252,129,129,0.3) !important;
    border-radius: 10px !important;
}

/* Divider */
hr {
    border-color: rgba(255,255,255,0.07) !important;
    margin: 1.5rem 0 !important;
}

/* Texto en general */
p, span, div, label {
    color: rgba(255,255,255,0.8);
}

h2, h3 {
    color: #e2e8f0 !important;
    font-family: 'Sora', sans-serif !important;
}

/* Badge de estado */
.badge {
    display: inline-block;
    padding: 0.2rem 0.7rem;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.04em;
}
.badge-blue { background: rgba(99,179,237,0.15); color: #90cdf4; border: 1px solid rgba(99,179,237,0.3); }
.badge-green { background: rgba(72,187,120,0.15); color: #9ae6b4; border: 1px solid rgba(72,187,120,0.3); }

/* Spinner */
.stSpinner > div {
    border-color: #63b3ed !important;
}

/* Ocultar footer de Streamlit */
footer { visibility: hidden; }
#MainMenu { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ============================================================
# ==================  UTILIDADES GLOBALES  ===================
# ============================================================

PATRONES_SEMESTRE = {
    'PRIMERO': [
        r'^B1(?!00)', r'^BS1(?!00)', r'^SB1(?!00)',
        r'^BR1(?!00)', r'^BRS1(?!00)', r'^1(?!0|1)',
        r'^C1(?!00)', r'^SC1(?!00)', r'^LB1(?!00)'
    ],
    'SEGUNDO': [
        r'^BS2', r'^SB2', r'^B2', r'^S2', r'^2'
    ],
    'TERCERO': [
        r'^BS3', r'^SB3', r'^B3', r'^S3', r'^3'
    ]
}

def determinar_semestre(grupo):
    if pd.isna(grupo):
        return None
    g = str(grupo).strip().upper()
    if not g:
        return None
    for semestre, patrones in PATRONES_SEMESTRE.items():
        if any(re.match(p, g) for p in patrones):
            return semestre
    return None

PATRONES_CALENDARIO = {
    'A': [
        r'^BS2', r'^SB2', r'^B2', r'^S2', r'^2',
        r'^BS3', r'^SB3', r'^B3', r'^S3', r'^3'
    ],
    'B': [
        r'^B1(?!00)', r'^BS1(?!00)', r'^SB1(?!00)',
        r'^BR1(?!00)', r'^BRS1(?!00)', r'^1(?!0|1)',
        r'^LB1(?!00)'
    ],
    'C': [
        r'^C1(?!00)', r'^SC1(?!00)'
    ]
}

BADGE_CALENDARIO_A = "BS2 · BS3 · SB2 · SB3 · B2 · B3 · S2 · S3 · 2 · 3"
BADGE_CALENDARIO_B = "B1 · BS1 · SB1 · BR1 · BRS1 · LB1 · 1"
BADGE_CALENDARIO_C = "C1 · SC1"
HELP_CALENDARIO_A = "Incluye grupos: BS2, BS3, SB2, SB3, B2, B3, S2, S3, 2, 3"
HELP_CALENDARIO_B = "Incluye grupos: B1, BS1, SB1, BR1, BRS1, LB1 (sin 00), 1 (sin 0 ni 1 después)"
HELP_CALENDARIO_C = "Incluye grupos: C1, SC1 (sin 00)"


def filtrar_grupo(grupo, calendario_a=True, calendario_b=True, calendario_c=True):
    if pd.isna(grupo):
        return False
    g = str(grupo).strip().upper()
    if not g:
        return False
    if calendario_a and any(re.match(p, g) for p in PATRONES_CALENDARIO['A']):
        return True
    if calendario_b and any(re.match(p, g) for p in PATRONES_CALENDARIO['B']):
        return True
    if calendario_c and any(re.match(p, g) for p in PATRONES_CALENDARIO['C']):
        return True
    return False


# ============================================================
# ==================  MÓDULO 1: FACULTADES  ==================
# ============================================================

MAPEO_COLUMNAS_FAC = {
    'DOCUMENTO': ['DOC', 'DOCUM', 'DOCUMENTO', 'PEGE_DOCUMENTOIDENTIDAD'],
    'NOMBRE': ['NOMBRE', 'NOM', 'ESTUDIANTE'],
    'SEMESTRE': ['SEM', 'SEMESTRE'],
    'PROGRAMA': ['PROGRAMA', 'PROG', 'PROG_NOMBRE'],
    'FACULTAD': ['FACULTAD'],
    'MATERIA': ['MATERIA'],
    'EVALUACION': ['EVALUACION'],
    'GRUPO': ['GRUPO'],
    'NOTA': ['NOTA'],
    'N_PERDIDAS': ['N_PERDIDAS']
}

CORRECCIONES_ORTOGRAFICAS = {
    'ADMINISTRACION': 'ADMINISTRACIÓN', 'ALGEBRA': 'ÁLGEBRA', 'ANATOMIA': 'ANATOMÍA',
    'AUDITORIA': 'AUDITORÍA', 'BASICAS': 'BÁSICAS', 'BIOLOGIA': 'BIOLOGÍA',
    'CALCULO': 'CÁLCULO', 'CIUDADANIA': 'CIUDADANÍA', 'COMUNICACION': 'COMUNICACIÓN',
    'CONSTITUCION': 'CONSTITUCIÓN', 'CONTADURIA': 'CONTADURÍA', 'ECONOMIA': 'ECONOMÍA',
    'ECOLOGIA': 'ECOLOGÍA', 'EDUCACION': 'EDUCACIÓN', 'ELECTRONICA': 'ELECTRÓNICA',
    'ESTADISTICA': 'ESTADÍSTICA', 'ETICA': 'ÉTICA', 'FISICA': 'FÍSICA',
    'FISIOLOGIA': 'FISIOLOGÍA', 'GESTION': 'GESTIÓN', 'INFORMACION': 'INFORMACIÓN',
    'INGENIERIA': 'INGENIERÍA', 'INGENIERIAS': 'INGENIERÍA', 'LEGISLACION': 'LEGISLACIÓN',
    'LOGICA': 'LÓGICA', 'MATEMATICA': 'MATEMÁTICA', 'MATEMATICAS': 'MATEMÁTICAS',
    'MECATRONICA': 'MECATRÓNICA', 'ORGANIZACION': 'ORGANIZACIÓN', 'PLANEACION': 'PLANEACIÓN',
    'POLITICA': 'POLÍTICA', 'PRODUCCION': 'PRODUCCIÓN', 'QUIMICA': 'QUÍMICA',
    'SOCIOLOGIA': 'SOCIOLOGÍA', 'EVALUACION': 'EVALUACIÓN', 'TECNOLOGIA': 'TECNOLOGÍA',
    'TEORIA': 'TEORÍA'
}

def fac_encontrar_columna(df, nombres_posibles):
    for nombre in nombres_posibles:
        nombre_norm = nombre.upper().translate(TILDES_MAP)
        for col in df.columns:
            if str(col).strip().upper().translate(TILDES_MAP) == nombre_norm:
                return col
    return None

def fac_normalizar_nota(nota):
    if pd.isna(nota):
        return nota
    try:
        return float(str(nota).replace(',', '.'))
    except:
        return nota

def fac_leer_excel(file_obj):
    try:
        validar_archivo_xlsx(file_obj)
        excel_file = pd.ExcelFile(file_obj, engine='openpyxl')
        sheet_names = [name.upper() for name in excel_file.sheet_names]
        if 'ORIGINAL' in sheet_names:
            idx = sheet_names.index('ORIGINAL')
            sheet_target = excel_file.sheet_names[idx]
        else:
            sheet_target = excel_file.sheet_names[0]

        # Intentar fila 2 primero (header=1), luego fila 1 (header=0)
        for header_row in [1, 0]:
            df = pd.read_excel(excel_file, sheet_name=sheet_target, header=header_row)
            cols_upper = [str(c).strip().upper() for c in df.columns]
            columnas_clave = ['DOCUMENTO', 'NOMBRE', 'PROGRAMA', 'FACULTAD', 'MATERIA', 'NOTA',
                              'DOC', 'DOCUM', 'PEGE_DOCUMENTOIDENTIDAD', 'PROG', 'PROG_NOMBRE']
            if any(c in cols_upper for c in columnas_clave):
                return df

        # Si ninguna funcionó, retornar con header=0 por defecto
        return pd.read_excel(excel_file, sheet_name=sheet_target, header=0)

    except Exception as e:
        raise Exception(f"Error al leer el archivo Excel: {str(e)}")

def fac_crear_hoja_general(df_original):
    df_general = pd.DataFrame()
    for col_norm, nombres in MAPEO_COLUMNAS_FAC.items():
        col_enc = fac_encontrar_columna(df_original, nombres)
        if col_enc:
            df_general[col_norm] = df_original[col_enc].copy()
    if 'SEMESTRE' not in df_general.columns:
        df_general['SEMESTRE'] = None
    if 'NOTA' in df_general.columns:
        df_general['NOTA'] = df_general['NOTA'].apply(fac_normalizar_nota)
    return df_general


def fac_calcular_n_perdidas(df):
    if 'DOCUMENTO' not in df.columns:
        return df
    df_resultado = df.copy()
    if 'N_PERDIDAS' not in df_resultado.columns:
        df_resultado['N_PERDIDAS'] = None
    conteo = df_resultado['DOCUMENTO'].value_counts().to_dict()
    df_resultado['N_PERDIDAS'] = df_resultado['DOCUMENTO'].map(conteo)
    df_resultado['N_PERDIDAS'] = df_resultado['N_PERDIDAS'].fillna(0).astype(int)
    return df_resultado

def fac_aplicar_filtros(df, calendario_a, calendario_b, calendario_c):
    df_f = df.copy()
    if 'NOTA' in df_f.columns:
        mask_nota = (df_f['NOTA'] >= 0.0) & (df_f['NOTA'] <= 2.9)
        df_f = df_f[mask_nota].copy()
    if 'GRUPO' in df_f.columns and (calendario_a or calendario_b or calendario_c):
        mask = df_f['GRUPO'].apply(lambda g: filtrar_grupo(g, calendario_a, calendario_b, calendario_c))
        df_f = df_f[mask].copy()
    if 'GRUPO' in df_f.columns and 'SEMESTRE' in df_f.columns:
        df_f['SEMESTRE'] = df_f['GRUPO'].apply(determinar_semestre)
    return df_f
    
def fac_corregir_ortografia_celda(valor):
    if valor is None or not isinstance(valor, str):
        return valor
    resultado = re.sub(r'(?i)^FACULTAD\s+DE\s+', '', valor).strip()
    for sin_tilde, con_tilde in CORRECCIONES_ORTOGRAFICAS.items():
        resultado = re.sub(r'\b' + sin_tilde + r'\b', con_tilde, resultado, flags=re.IGNORECASE)
    return resultado

def fac_corregir_ortografia_hoja(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = fac_corregir_ortografia_celda(cell.value)

def fac_aplicar_formato_especial(workbook, nombre_hoja, titulo_encabezado):
    if nombre_hoja not in workbook.sheetnames:
        return
    sheet = workbook[nombre_hoja]
    fac_corregir_ortografia_hoja(sheet)
    if sheet.max_row < 4:
        return
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        if cell.value is not None:
            cell.value = None
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1)
        if cell.value is not None:
            cell.value = None
    ultima_columna_real = 1
    for col in range(2, sheet.max_column + 1):
        cell = sheet.cell(row=4, column=col)
        if cell.value is not None and str(cell.value).strip() != '':
            ultima_columna_real = col
    if ultima_columna_real == 1 and sheet.max_row >= 5:
        for col in range(2, sheet.max_column + 1):
            cell = sheet.cell(row=5, column=col)
            if cell.value is not None:
                ultima_columna_real = col
    if ultima_columna_real == 1:
        ultima_columna_real = sheet.max_column
    ultima_columna_letra = get_column_letter(ultima_columna_real)
    rango_encabezado = f'B2:{ultima_columna_letra}3'
    sheet.merge_cells(rango_encabezado)
    celda_encabezado = sheet['B2']
    celda_encabezado.value = titulo_encabezado
    celda_encabezado.font = Font(bold=True, size=20)
    celda_encabezado.alignment = Alignment(horizontal='center', vertical='center')
    columnas_nombres = {}
    for col in range(2, ultima_columna_real + 1):
        cell = sheet.cell(row=4, column=col)
        if cell.value is not None:
            columnas_nombres[str(cell.value).upper().strip()] = col
    sin_borde = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
    for col in range(2, ultima_columna_real + 1):
        cell = sheet.cell(row=4, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = sin_borde
    rango_filtro = f'B4:{ultima_columna_letra}{sheet.max_row}'
    sheet.auto_filter.ref = rango_filtro
    formato_numero = '0'
    if 'DOCUMENTO' in columnas_nombres:
        col_documento = columnas_nombres['DOCUMENTO']
        for row in range(5, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_documento)
            if cell.value is not None:
                try:
                    cell.value = float(str(cell.value).replace(',', '.'))
                    cell.number_format = formato_numero
                except:
                    pass
    if 'GRUPO' in columnas_nombres:
        col_grupo = columnas_nombres['GRUPO']
        for row in range(5, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_grupo)
            if cell.value is not None:
                cell.value = str(cell.value).strip()
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left', vertical='center')
    if 'N_PERDIDAS' in columnas_nombres:
        col_n_perdidas = columnas_nombres['N_PERDIDAS']
        for row in range(5, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_n_perdidas)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(2, ultima_columna_real + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 18

def fac_aplicar_formato_normal(workbook, nombre_hoja):
    if nombre_hoja not in workbook.sheetnames:
        return
    sheet = workbook[nombre_hoja]
    fac_corregir_ortografia_hoja(sheet)
    if sheet.max_row < 1:
        return
    columnas_nombres = {}
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        if cell.value is not None:
            columnas_nombres[str(cell.value).upper().strip()] = col
    sin_borde = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = sin_borde
    ultima_columna_letra = get_column_letter(sheet.max_column)
    rango_filtro = f'A1:{ultima_columna_letra}{sheet.max_row}'
    sheet.auto_filter.ref = rango_filtro
    formato_numero = '0'
    if 'DOCUMENTO' in columnas_nombres:
        col_documento = columnas_nombres['DOCUMENTO']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_documento)
            if cell.value is not None:
                try:
                    cell.value = float(str(cell.value).replace(',', '.'))
                    cell.number_format = formato_numero
                except:
                    pass
    if 'GRUPO' in columnas_nombres:
        col_grupo = columnas_nombres['GRUPO']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_grupo)
            if cell.value is not None:
                cell.value = str(cell.value).strip()
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left', vertical='center')
    if 'N_PERDIDAS' in columnas_nombres:
        col_n_perdidas = columnas_nombres['N_PERDIDAS']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_n_perdidas)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 18


def fac_limpiar_nombre_hoja(nombre, nombres_usados):
    nombre_limpio = str(nombre)[:31]
    for ch in ['/', '\\', '?', '*', '[', ']', ':']:
        nombre_limpio = nombre_limpio.replace(ch, '_')
    nombre_original = nombre_limpio
    contador = 1
    while nombre_limpio in nombres_usados:
        nombre_limpio = f"{nombre_original[:27]}_{contador}"
        contador += 1
    nombres_usados.add(nombre_limpio)
    return nombre_limpio


def fac_procesar_archivo(file_obj1, file_obj2, calendario_a, calendario_b, calendario_c):
    """Procesa los archivos de facultades y retorna un dict {nombre_archivo: bytes}"""

    # Leer y unir los dos archivos (el segundo es opcional)
    df1 = fac_leer_excel(file_obj1)
    if file_obj2 is not None:
        df2 = fac_leer_excel(file_obj2)
        df_original = pd.concat([df1, df2], ignore_index=True)
    else:
        df_original = df1

    df_general = fac_crear_hoja_general(df_original)
    if df_general.empty:
        raise Exception("No se pudieron mapear las columnas del archivo original.")
    if 'FACULTAD' not in df_general.columns:
        raise Exception("No se encontró la columna FACULTAD en el archivo.")

    # Limpiar columna PROGRAMA: eliminar " - ..." desde el guion hacia adelante
    if 'PROGRAMA' in df_general.columns:
        df_general['PROGRAMA'] = df_general['PROGRAMA'].apply(
            lambda x: str(x).split(' - ')[0].strip() if pd.notna(x) else x
        )

    facultades = df_general['FACULTAD'].dropna().unique()
    if len(facultades) == 0:
        raise Exception("No se encontraron facultades en el archivo.")

    archivos_resultado = {}
    todas_hojas_general = []

    for facultad in facultades:
        df_facultad = df_general[df_general['FACULTAD'] == facultad].copy()
        df_facultad = fac_aplicar_filtros(df_facultad, calendario_a, calendario_b, calendario_c)
        programas = df_facultad['PROGRAMA'].dropna().unique() if 'PROGRAMA' in df_facultad.columns else []

        nombre_facultad_limpio = "".join(c for c in str(facultad) if c.isalnum() or c in (' ', '-', '_')).strip()
        nombre_facultad_limpio = nombre_facultad_limpio.replace(' ', '_')

        nombres_hojas_usados = set(['GENERAL'])
        programa_a_nombre_hoja = {}
        for programa in programas:
            nombre_hoja = fac_limpiar_nombre_hoja(programa, nombres_hojas_usados)
            programa_a_nombre_hoja[programa] = nombre_hoja

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            columnas_ordenadas = ['DOCUMENTO', 'NOMBRE', 'SEMESTRE', 'PROGRAMA', 'MATERIA', 'GRUPO', 'NOTA', 'N_PERDIDAS']
            columnas_existentes = [col for col in columnas_ordenadas if col in df_facultad.columns]
            df_general_facultad = df_facultad[columnas_existentes].copy()
            df_general_facultad = fac_calcular_n_perdidas(df_general_facultad)
            df_para_informe = df_general_facultad.copy()
            df_para_informe['FACULTAD'] = facultad
            if 'EVALUACION' in df_facultad.columns:
                df_para_informe['EVALUACION'] = df_facultad['EVALUACION'].values
            todas_hojas_general.append(df_para_informe)
            df_general_facultad.to_excel(writer, sheet_name='GENERAL', index=False, startrow=3, startcol=1)
            for programa in programas:
                df_programa = df_facultad[df_facultad['PROGRAMA'] == programa].copy()
                df_programa = df_programa[columnas_existentes].copy()
                if 'DOCUMENTO' in df_general_facultad.columns and 'N_PERDIDAS' in df_general_facultad.columns:
                    doc_to_np = dict(zip(df_general_facultad['DOCUMENTO'], df_general_facultad['N_PERDIDAS']))
                    if 'DOCUMENTO' in df_programa.columns:
                        df_programa['N_PERDIDAS'] = df_programa['DOCUMENTO'].map(doc_to_np)
                        df_programa['N_PERDIDAS'] = df_programa['N_PERDIDAS'].fillna(0).astype(int)
                nombre_hoja = programa_a_nombre_hoja[programa]
                df_programa.to_excel(writer, sheet_name=nombre_hoja, index=False, startrow=3, startcol=1)

        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        titulo_fac = fac_corregir_ortografia_celda(f'REPORTE GENERAL {facultad}')
        fac_aplicar_formato_especial(workbook, 'GENERAL', titulo_fac)
        for programa in programas:
            nombre_hoja = programa_a_nombre_hoja[programa]
            if nombre_hoja in workbook.sheetnames:
                fac_aplicar_formato_especial(workbook, nombre_hoja, fac_corregir_ortografia_celda(str(programa)))
        final_output = io.BytesIO()
        workbook.save(final_output)
        final_output.seek(0)
        archivos_resultado[f"{nombre_facultad_limpio}.xlsx"] = final_output.getvalue()

    # Crear INFORME GENERAL
    if todas_hojas_general:
        df_informe_general = pd.concat(todas_hojas_general, ignore_index=True)
        df_informe_general = fac_calcular_n_perdidas(df_informe_general)
        columnas_ordenadas_ig = ['DOCUMENTO', 'NOMBRE', 'SEMESTRE', 'PROGRAMA', 'FACULTAD', 'MATERIA', 'EVALUACION', 'GRUPO', 'NOTA', 'N_PERDIDAS']
        columnas_existentes_ig = [col for col in columnas_ordenadas_ig if col in df_informe_general.columns]
        df_informe_general = df_informe_general[columnas_existentes_ig].copy()
        output_ig = io.BytesIO()
        with pd.ExcelWriter(output_ig, engine='openpyxl') as writer:
            df_informe_general.to_excel(writer, sheet_name='GENERAL', index=False, startrow=0, startcol=0)
        output_ig.seek(0)
        wb_ig = openpyxl.load_workbook(output_ig)
        fac_aplicar_formato_normal(wb_ig, 'GENERAL')
        final_ig = io.BytesIO()
        wb_ig.save(final_ig)
        final_ig.seek(0)
        archivos_resultado["INFORME GENERAL.xlsx"] = final_ig.getvalue()

    return archivos_resultado

# ============================================================
# ==================  MÓDULO 2: MATRICULADOS  ================
# ============================================================

MAPEO_COLUMNAS_MAT = {
    'DOCUMENTO': ['DOC', 'DOCUM', 'DOCUMENTO', 'PEGE_DOCUMENTOIDENTIDAD'],
    'NOMBRE': ['NOMBRE', 'NOM', 'ESTUDIANTE'],
    'PROGRAMA': ['PROGRAMA', 'PROG', 'PROG_NOMBRE'],
    'SEDE': ['FRANJA', 'SEDE'],
    'MATERIA': ['MATERIA'],
    'GRUPO': ['GRUPO'],
    'SEMESTRE': ['SEMESTRE'],
    'CELULAR': ['CELULAR'],
    'CORREO_INST': ['CORREO_INST']
}

PROGRAMA_A_FACULTAD = {
    'ADMINISTRACION DE EMPRESAS': 'CIENCIAS EMPRESARIALES',
    'CONTADURIA PUBLICA': 'CIENCIAS EMPRESARIALES',
    'TECNOLOGIA EN CONTABILIDAD SISTEMATIZADA': 'CIENCIAS EMPRESARIALES',
    'TECNOLOGIA EN GESTION CONTABLE Y FINANCIERA': 'CIENCIAS EMPRESARIALES',
    'TECNOLOGIA EN GESTION EMPRESARIAL': 'CIENCIAS EMPRESARIALES',
    'TECNOLOGIA EN GESTION LOGISTICA': 'CIENCIAS EMPRESARIALES',
    'TECNOLOGIA EN MERCADEO Y NEGOCIOS INTERNACIONALES': 'CIENCIAS EMPRESARIALES',
    'COMUNICACION SOCIAL': 'CIENCIAS SOCIALES Y HUMANAS',
    'DISENO VISUAL': 'CIENCIAS SOCIALES Y HUMANAS',
    'TRABAJO SOCIAL': 'CIENCIAS SOCIALES Y HUMANAS',
    'ADMINISTRACION EN SALUD': 'EDUCACION A DISTANCIA Y VIRTUAL',
    'ADMINISTRACION EN SEGURIDAD Y SALUD EN EL TRABAJO': 'EDUCACION A DISTANCIA Y VIRTUAL',
    'LICENCIATURA EN CIENCIAS DEL DEPORTE Y LA EDUCACION FISICA': 'EDUCACION A DISTANCIA Y VIRTUAL',
    'LICENCIATURA EN EDUCACION INFANTIL': 'EDUCACION A DISTANCIA Y VIRTUAL',
    'SALUD OCUPACIONAL': 'EDUCACION A DISTANCIA Y VIRTUAL',
    'INGENIERIA DE SISTEMAS': 'INGENIERIA',
    'INGENIERIA ELECTRONICA': 'INGENIERIA',
    'INGENIERIA INDUSTRIAL': 'INGENIERIA',
    'TECNOLOGIA EN ELECTRONICA INDUSTRIAL': 'INGENIERIA',
    'TECNOLOGIA EN MECATRONICA INDUSTRIAL': 'INGENIERIA',
    'TECNOLOGIA EN PRODUCCION INDUSTRIAL': 'INGENIERIA',
    'TECNOLOGIA EN SISTEMAS DE INFORMACION': 'INGENIERIA',
}

TILDES_MAP = str.maketrans('ÁÉÍÓÚáéíóúÑñÜü', 'AEIOUaeiouNnUu')

def mat_limpiar_texto(texto):
    if pd.isna(texto):
        return texto
    return str(texto).translate(TILDES_MAP)

def mat_normalizar_programa(programa):
    if pd.isna(programa):
        return programa
    return mat_limpiar_texto(str(programa).strip().upper())

MATERIAS_COMUNICACION_MAT = [
    'COMUNICACION Y LENGUAJE', 'COMUNICACION Y LENGUAJE I', 'COMUNICACION Y LENGUAJE II',
    'COMUNICACION', 'COMUNICACION I', 'COMUNICACION II',
    'SEMINARIO DE COMUNICACION', 'SEMINARIO DE COMUNICACION I', 'SEMINARIO DE COMUNICACION II'
]

MATERIAS_MATEMATICAS_MAT = [
    'ALGEBRA LINEAL', 'MATEMATICAS', 'MATEMATICAS I', 'MATEMATICAS II', 'MATEMATICAS III',
    'MATEMATICAS FUNDAMENTAL', 'MATEMATICAS FUNDAMENTALES', 'MATEMATICA FUNDAMENTAL',
    'MATEMATICAS BASICAS', 'LOGICA Y RAZONAMIENTO', 'CALCULO INTEGRAL', 'CALCULO DIFERENCIAL', 'CALCULO'
]

def mat_encontrar_columna(df, nombres_posibles):
    for nombre in nombres_posibles:
        nombre_norm = nombre.upper().translate(TILDES_MAP)
        for col in df.columns:
            if str(col).strip().upper().translate(TILDES_MAP) == nombre_norm:
                return col
    return None

def mat_normalizar_sede(sede):
    if pd.isna(sede):
        return sede
    sede_str = str(sede).upper()
    if 'NORTE' in sede_str: return 'NORTE'
    elif 'SUR' in sede_str: return 'SUR'
    return sede



def mat_leer_excel(file_obj):
    try:
        validar_archivo_xlsx(file_obj)
        excel_file = pd.ExcelFile(file_obj, engine='openpyxl')
        sheet_names = [name.upper() for name in excel_file.sheet_names]
        if 'ORIGINAL' in sheet_names:
            idx = sheet_names.index('ORIGINAL')
            df = pd.read_excel(excel_file, sheet_name=excel_file.sheet_names[idx])
        else:
            df = pd.read_excel(excel_file, sheet_name=0)
        return df
    except Exception as e:
        raise Exception(f"Error al leer el archivo Excel: {str(e)}")

def mat_crear_hoja_general(df_original):
    df_general = pd.DataFrame()
    for col_norm, nombres in MAPEO_COLUMNAS_MAT.items():
        col_enc = mat_encontrar_columna(df_original, nombres)
        if col_enc:
            df_general[col_norm] = df_original[col_enc].copy()
    if 'PROGRAMA' in df_general.columns:
        df_general['FACULTAD'] = df_general['PROGRAMA'].apply(
            lambda p: PROGRAMA_A_FACULTAD.get(mat_normalizar_programa(p), None)
        )
        cols = list(df_general.columns)
        if 'FACULTAD' in cols and 'PROGRAMA' in cols:
            cols.remove('FACULTAD')
            idx_prog = cols.index('PROGRAMA')
            cols.insert(idx_prog, 'FACULTAD')
            df_general = df_general[cols]
    if 'SEDE' in df_general.columns:
        df_general['SEDE'] = df_general['SEDE'].apply(mat_normalizar_sede)
    if 'GRUPO' in df_general.columns:
        mask = df_general['GRUPO'].apply(lambda g: filtrar_grupo(g, calendario_a=True, calendario_b=True, calendario_c=True))
        df_general = df_general[mask].copy()
    if 'GRUPO' in df_general.columns:
        df_general['SEMESTRE'] = df_general['GRUPO'].apply(determinar_semestre)
    return df_general

def mat_crear_hoja_matriculados(df_general):
    if 'GRUPO' not in df_general.columns or 'DOCUMENTO' not in df_general.columns:
        return pd.DataFrame(columns=['GRUPO', 'MATRICULADOS'])
    df_mat = df_general.groupby('GRUPO')['DOCUMENTO'].nunique().reset_index()
    df_mat.columns = ['GRUPO', 'MATRICULADOS']
    df_mat = df_mat.sort_values('GRUPO').reset_index(drop=True)
    return df_mat

def mat_filtrar_por_materias(df_general, materias):
    if 'MATERIA' not in df_general.columns:
        return pd.DataFrame()
    df_normalizado = df_general['MATERIA'].astype(str).str.upper().str.strip().str.translate(TILDES_MAP)
    materias_normalizado = [str(m).upper().strip().translate(TILDES_MAP) for m in materias]
    mask = df_normalizado.isin(materias_normalizado)
    return df_general[mask].copy()

def mat_aplicar_formato(workbook, nombre_hoja):
    if nombre_hoja not in workbook.sheetnames:
        return
    sheet = workbook[nombre_hoja]
    columnas_nombres = {}
    if sheet.max_row > 0:
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value is not None:
                columnas_nombres[str(cell.value).upper().strip()] = idx
    if sheet.max_row > 0:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    if sheet.max_row > 0:
        sheet.auto_filter.ref = sheet.dimensions
    formato_numero = '0'
    if 'DOCUMENTO' in columnas_nombres and sheet.max_row > 1:
        col_documento = columnas_nombres['DOCUMENTO']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_documento)
            if cell.value is not None:
                try:
                    cell.value = float(str(cell.value).replace(',', '.'))
                    cell.number_format = formato_numero
                except:
                    pass
    if 'GRUPO' in columnas_nombres and sheet.max_row > 1:
        col_grupo = columnas_nombres['GRUPO']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_grupo)
            if cell.value is not None:
                cell.value = str(cell.value).strip()
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left', vertical='center')
    if 'N_PERDIDAS' in columnas_nombres and sheet.max_row > 1:
        col_n_perdidas = columnas_nombres['N_PERDIDAS']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_n_perdidas)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    if 'CELULAR' in columnas_nombres and sheet.max_row > 1:
        col_celular = columnas_nombres['CELULAR']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_celular)
            if cell.value is not None:
                try:
                    valor_limpio = str(cell.value).replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                    cell.value = float(valor_limpio.replace(',', '.'))
                    cell.number_format = formato_numero
                except:
                    pass
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 18

def mat_procesar_archivo(file_obj):
    df_original = mat_leer_excel(file_obj)
    df_general = mat_crear_hoja_general(df_original)
    if df_general.empty:
        raise Exception("No se pudieron mapear las columnas del archivo original o no hay datos después del filtrado.")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_general.to_excel(writer, sheet_name='GENERAL', index=False)
        df_matriculados = mat_crear_hoja_matriculados(df_general)
        if not df_matriculados.empty:
            df_matriculados.to_excel(writer, sheet_name='MATRICULADOS', index=False)
        else:
            pd.DataFrame(columns=['GRUPO', 'MATRICULADOS']).to_excel(writer, sheet_name='MATRICULADOS', index=False)
        df_comunicacion = mat_filtrar_por_materias(df_general, MATERIAS_COMUNICACION_MAT)
        if not df_comunicacion.empty:
            df_comunicacion.to_excel(writer, sheet_name='COMUNICACION', index=False)
        else:
            df_general.head(0).to_excel(writer, sheet_name='COMUNICACION', index=False)
        df_matematicas = mat_filtrar_por_materias(df_general, MATERIAS_MATEMATICAS_MAT)
        if not df_matematicas.empty:
            df_matematicas.to_excel(writer, sheet_name='MATEMATICAS', index=False)
        else:
            df_general.head(0).to_excel(writer, sheet_name='MATEMATICAS', index=False)
    output.seek(0)
    workbook = openpyxl.load_workbook(output)
    mat_aplicar_formato(workbook, 'GENERAL')
    mat_aplicar_formato(workbook, 'MATRICULADOS')
    mat_aplicar_formato(workbook, 'COMUNICACION')
    mat_aplicar_formato(workbook, 'MATEMATICAS')
    final_output = io.BytesIO()
    workbook.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()


# ============================================================
# ==================  MÓDULO 3: CRUCE  =======================
# ============================================================

MAPEO_COLUMNAS_CRUCE = {
    'DOCUMENTO': ['DOC', 'DOCUM', 'DOCUMENTO', 'PEGE_DOCUMENTOIDENTIDAD'],
    'NOMBRE': ['NOMBRE', 'NOM', 'ESTUDIANTE'],
    'FACULTAD': ['FACULTAD'],
    'PROGRAMA': ['PROGRAMA', 'PROG', 'PROG_NOMBRE'],
    'SEDE': ['FRANJA', 'SEDE'],
    'TEMA': ['ENCU_TEMA', 'TEMA'],
    'FECHA': ['FECHA_ENCUENTRO', 'FECHA'],
    'AREA': ['AREA'],
    'TUTOR': ['TUTOR', 'INSTRUCTOR']
}

def cruce_encontrar_columna(df, nombres_posibles):
    for nombre in nombres_posibles:
        nombre_norm = nombre.upper().translate(TILDES_MAP)
        for col in df.columns:
            if str(col).strip().upper().translate(TILDES_MAP) == nombre_norm:
                return col
    return None

def cruce_leer_excel(file_obj, nombre_hoja=None):
    try:
        validar_archivo_xlsx(file_obj)
        excel_file = pd.ExcelFile(file_obj, engine='openpyxl')
        if nombre_hoja:
            sheet_names = [name.upper() for name in excel_file.sheet_names]
            nombre_hoja_upper = nombre_hoja.upper()
            if nombre_hoja_upper in sheet_names:
                idx = sheet_names.index(nombre_hoja_upper)
                df = pd.read_excel(excel_file, sheet_name=excel_file.sheet_names[idx])
            else:
                raise Exception(f"No se encontró la hoja '{nombre_hoja}' en el archivo.")
        else:
            sheet_names = [name.upper() for name in excel_file.sheet_names]
            if 'ORIGINAL' in sheet_names:
                idx = sheet_names.index('ORIGINAL')
                df = pd.read_excel(excel_file, sheet_name=excel_file.sheet_names[idx])
            else:
                df = pd.read_excel(excel_file, sheet_name=0)
        return df
    except Exception as e:
        raise Exception(f"Error al leer el archivo Excel: {str(e)}")

def cruce_crear_hoja_general(df_original):
    df_general = pd.DataFrame()
    for col_norm, nombres in MAPEO_COLUMNAS_CRUCE.items():
        col_enc = cruce_encontrar_columna(df_original, nombres)
        if col_enc:
            df_general[col_norm] = df_original[col_enc].copy()
    if 'FECHA' in df_general.columns:
        df_general['FECHA'] = pd.to_datetime(df_general['FECHA'], errors='coerce').dt.date
    return df_general

def cruce_filtrar_por_area(df_general, valores_area):
    if 'AREA' not in df_general.columns:
        return pd.DataFrame()
    df_area_upper = df_general['AREA'].astype(str).str.upper().str.strip()
    valores_upper = [v.upper().strip() for v in valores_area]
    mask = df_area_upper.isin(valores_upper)
    df_filtrado = df_general[mask].copy()
    if 'DOCUMENTO' in df_filtrado.columns:
        df_filtrado = df_filtrado.drop_duplicates(subset=['DOCUMENTO'], keep='first')
    return df_filtrado

def cruce_aplicar_filtro_grupos(df, calendario_a, calendario_b, calendario_c):
    if 'GRUPO' not in df.columns: return df
    if not calendario_a and not calendario_b and not calendario_c: return df
    mask = df['GRUPO'].apply(lambda g: filtrar_grupo(g, calendario_a, calendario_b, calendario_c))
    return df[mask].copy()

def cruce_cruzar_con_matriculados(df_asistencia, file_obj_matriculados, hoja_matriculados):
    try:
        df_matriculados = cruce_leer_excel(file_obj_matriculados, hoja_matriculados)
        if 'DOCUMENTO' not in df_asistencia.columns:
            return df_asistencia
        if 'DOCUMENTO' not in df_matriculados.columns:
            return pd.DataFrame(columns=df_asistencia.columns)
        doc_to_grupo = {}
        doc_to_sede = {}
        if 'GRUPO' in df_matriculados.columns:
            for idx, row in df_matriculados.iterrows():
                if pd.notna(row.get('DOCUMENTO')):
                    doc = str(row['DOCUMENTO']).strip()
                    if pd.notna(row.get('GRUPO')):
                        doc_to_grupo[doc] = row['GRUPO']
        if 'SEDE' in df_matriculados.columns:
            for idx, row in df_matriculados.iterrows():
                if pd.notna(row.get('DOCUMENTO')):
                    doc = str(row['DOCUMENTO']).strip()
                    if pd.notna(row.get('SEDE')):
                        doc_to_sede[doc] = row['SEDE']
        df_resultado = df_asistencia.copy()
        if 'TUTOR' in df_resultado.columns:
            tutor_idx = list(df_resultado.columns).index('TUTOR')
            df_resultado.insert(tutor_idx + 1, 'GRUPO', None)
        else:
            df_resultado['GRUPO'] = None
        indices_a_eliminar = []
        for idx, row in df_resultado.iterrows():
            if pd.notna(row.get('DOCUMENTO')):
                doc = str(row['DOCUMENTO']).strip()
                if doc in doc_to_grupo:
                    df_resultado.at[idx, 'GRUPO'] = doc_to_grupo[doc]
                else:
                    indices_a_eliminar.append(idx)
                    continue
                if doc in doc_to_sede:
                    df_resultado.at[idx, 'SEDE'] = doc_to_sede[doc]
            else:
                indices_a_eliminar.append(idx)
        df_resultado = df_resultado.drop(indices_a_eliminar)
        df_resultado = df_resultado.reset_index(drop=True)
        return df_resultado
    except Exception as e:
        raise Exception(f"Error al cruzar con MATRICULADOS: {str(e)}")

def cruce_actualizar_sede_desde_matriculados(df_asistencia, file_obj_matriculados, hoja_matriculados):
    try:
        df_matriculados = cruce_leer_excel(file_obj_matriculados, hoja_matriculados)
        if 'DOCUMENTO' not in df_asistencia.columns or 'DOCUMENTO' not in df_matriculados.columns:
            return df_asistencia
        if 'SEDE' not in df_matriculados.columns:
            return df_asistencia
        doc_to_sede = {}
        for idx, row in df_matriculados.iterrows():
            if pd.notna(row.get('DOCUMENTO')):
                doc = str(row['DOCUMENTO']).strip()
                if pd.notna(row.get('SEDE')):
                    doc_to_sede[doc] = row['SEDE']
        df_resultado = df_asistencia.copy()
        for idx, row in df_resultado.iterrows():
            if pd.notna(row.get('DOCUMENTO')):
                doc = str(row['DOCUMENTO']).strip()
                if doc in doc_to_sede:
                    df_resultado.at[idx, 'SEDE'] = doc_to_sede[doc]
        return df_resultado
    except Exception as e:
        raise Exception(f"Error al actualizar SEDE desde MATRICULADOS: {str(e)}")

def cruce_crear_hoja_asistencia_pma(df_matematicas, df_comunicacion):
    df_union = pd.concat([df_matematicas, df_comunicacion], ignore_index=True)
    if 'DOCUMENTO' in df_union.columns and 'GRUPO' in df_union.columns:
        df_union = df_union.drop_duplicates(subset=['DOCUMENTO', 'GRUPO'], keep='first')
    return df_union.reset_index(drop=True)

def cruce_filtrar_ganaron(df_asistencia, file_obj_informe, hoja_informe):
    try:
        df_informe = cruce_leer_excel(file_obj_informe, hoja_informe)
        if 'DOCUMENTO' not in df_asistencia.columns:
            return df_asistencia
        if 'DOCUMENTO' not in df_informe.columns:
            return df_asistencia
        docs_perdidos = set()
        for idx, row in df_informe.iterrows():
            if pd.notna(row.get('DOCUMENTO')):
                doc = str(row['DOCUMENTO']).strip()
                docs_perdidos.add(doc)
        df_resultado = df_asistencia.copy()
        mask = df_resultado['DOCUMENTO'].apply(
            lambda x: pd.notna(x) and str(x).strip() not in docs_perdidos
        )
        return df_resultado[mask].reset_index(drop=True)
    except Exception as e:
        raise Exception(f"Error al filtrar ganaron: {str(e)}")

def cruce_eliminar_columna_semestre(df):
    if 'SEMESTRE' in df.columns:
        df = df.drop(columns=['SEMESTRE'])
    return df

def cruce_aplicar_formato(workbook, nombre_hoja):
    if nombre_hoja not in workbook.sheetnames:
        return
    sheet = workbook[nombre_hoja]
    columnas_nombres = {}
    if sheet.max_row > 0:
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value is not None:
                columnas_nombres[str(cell.value).upper().strip()] = idx
    if sheet.max_row > 0:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    if sheet.max_row > 0:
        sheet.auto_filter.ref = sheet.dimensions
    formato_numero = '0'
    if 'DOCUMENTO' in columnas_nombres and sheet.max_row > 1:
        col_documento = columnas_nombres['DOCUMENTO']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_documento)
            if cell.value is not None:
                try:
                    cell.value = float(str(cell.value).replace(',', '.'))
                    cell.number_format = formato_numero
                except:
                    pass
    if 'GRUPO' in columnas_nombres and sheet.max_row > 1:
        col_grupo = columnas_nombres['GRUPO']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_grupo)
            if cell.value is not None:
                cell.value = str(cell.value).strip()
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left', vertical='center')
    if 'N_PERDIDAS' in columnas_nombres and sheet.max_row > 1:
        col_n_perdidas = columnas_nombres['N_PERDIDAS']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_n_perdidas)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 18

def cruce_procesar_archivo(file_cruce, file_matriculados, file_informe, calendario_a, calendario_b, calendario_c):
    df_original = cruce_leer_excel(file_cruce)
    df_general = cruce_crear_hoja_general(df_original)
    df_general = cruce_eliminar_columna_semestre(df_general)
    if df_general.empty:
        raise Exception("No se pudieron mapear las columnas del archivo CRUCE.")

    df_asistencia_matematicas = cruce_filtrar_por_area(df_general, ['M', 'DCB'])
    if not df_asistencia_matematicas.empty:
        file_matriculados.seek(0)
        df_asistencia_matematicas = cruce_cruzar_con_matriculados(df_asistencia_matematicas, file_matriculados, 'MATEMATICAS')
        if not df_asistencia_matematicas.empty:
            file_matriculados.seek(0)
            df_asistencia_matematicas = cruce_actualizar_sede_desde_matriculados(df_asistencia_matematicas, file_matriculados, 'MATEMATICAS')
            df_asistencia_matematicas = cruce_eliminar_columna_semestre(df_asistencia_matematicas)

    df_asistencia_comunicacion = cruce_filtrar_por_area(df_general, ['C'])
    if not df_asistencia_comunicacion.empty:
        file_matriculados.seek(0)
        df_asistencia_comunicacion = cruce_cruzar_con_matriculados(df_asistencia_comunicacion, file_matriculados, 'COMUNICACION')
        if not df_asistencia_comunicacion.empty:
            file_matriculados.seek(0)
            df_asistencia_comunicacion = cruce_actualizar_sede_desde_matriculados(df_asistencia_comunicacion, file_matriculados, 'COMUNICACION')
            df_asistencia_comunicacion = cruce_eliminar_columna_semestre(df_asistencia_comunicacion)

    if calendario_a or calendario_b or calendario_c:
        if not df_asistencia_matematicas.empty:
            df_asistencia_matematicas = cruce_aplicar_filtro_grupos(df_asistencia_matematicas, calendario_a, calendario_b, calendario_c)
        if not df_asistencia_comunicacion.empty:
            df_asistencia_comunicacion = cruce_aplicar_filtro_grupos(df_asistencia_comunicacion, calendario_a, calendario_b, calendario_c)

    df_asistencia_pma = cruce_crear_hoja_asistencia_pma(df_asistencia_matematicas, df_asistencia_comunicacion)
    df_asistencia_pma = cruce_eliminar_columna_semestre(df_asistencia_pma)

    file_informe.seek(0)
    df_ganaron_matematicas = cruce_filtrar_ganaron(df_asistencia_matematicas, file_informe, 'MATEMATICAS')
    df_ganaron_matematicas = cruce_eliminar_columna_semestre(df_ganaron_matematicas)

    file_informe.seek(0)
    df_ganaron_comunicacion = cruce_filtrar_ganaron(df_asistencia_comunicacion, file_informe, 'COMUNICACION')
    df_ganaron_comunicacion = cruce_eliminar_columna_semestre(df_ganaron_comunicacion)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_general.to_excel(writer, sheet_name='GENERAL', index=False)
        (df_asistencia_matematicas if not df_asistencia_matematicas.empty else df_general.head(0)).to_excel(writer, sheet_name='ASISTENCIA MATEMATICAS', index=False)
        (df_asistencia_comunicacion if not df_asistencia_comunicacion.empty else df_general.head(0)).to_excel(writer, sheet_name='ASISTENCIA COMUNICACION', index=False)
        (df_asistencia_pma if not df_asistencia_pma.empty else df_general.head(0)).to_excel(writer, sheet_name='ASISTENCIA PMA', index=False)
        (df_ganaron_matematicas if not df_ganaron_matematicas.empty else df_general.head(0)).to_excel(writer, sheet_name='GANARON MATEMATICAS', index=False)
        (df_ganaron_comunicacion if not df_ganaron_comunicacion.empty else df_general.head(0)).to_excel(writer, sheet_name='GANARON COMUNICACION', index=False)

    output.seek(0)
    workbook = openpyxl.load_workbook(output)
    for hoja in ['GENERAL', 'ASISTENCIA MATEMATICAS', 'ASISTENCIA COMUNICACION', 'ASISTENCIA PMA', 'GANARON MATEMATICAS', 'GANARON COMUNICACION']:
        cruce_aplicar_formato(workbook, hoja)
    final_output = io.BytesIO()
    workbook.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()


# ============================================================
# ==================  MÓDULO 4: INFORME  =====================
# ============================================================

MAPEO_COLUMNAS_INF = {
    'DOCUMENTO': ['DOC', 'DOCUM', 'DOCUMENTO', 'PEGE_DOCUMENTOIDENTIDAD'],
    'NOMBRE': ['NOMBRE', 'NOM', 'ESTUDIANTE'],
    'SEMESTRE': ['SEM', 'SEMESTRE'],
    'SEDE': ['FRANJA', 'SEDE'],
    'PROGRAMA': ['PROGRAMA', 'PROG', 'PROG_NOMBRE'],
    'FACULTAD': ['FACULTAD'],
    'MATERIA': ['MATERIA'],
    'EVALUACIÓN': ['EVALUACION'],
    'GRUPO': ['GRUPO'],
    'NOTA': ['NOTA'],
    'N_PERDIDAS': ['N_PERDIDAS']
}

MATERIAS_COMUNICACION_INF = [
    'COMUNICACIÓN Y LENGUAJE', 'COMUNICACIÓN Y LENGUAJE I', 'COMUNICACIÓN Y LENGUAJE II',
    'COMUNICACIÓN', 'COMUNICACIÓN I', 'COMUNICACIÓN II',
    'SEMINARIO DE COMUNICACIÓN', 'SEMINARIO DE COMUNICACIÓN I', 'SEMINARIO DE COMUNICACIÓN II'
]

MATERIAS_MATEMATICAS_INF = [
    'ÁLGEBRA LINEAL', 'MATEMÁTICAS', 'MATEMÁTICAS I', 'MATEMÁTICAS II', 'MATEMÁTICAS III',
    'MATEMÁTICAS FUNDAMENTAL', 'MATEMÁTICAS FUNDAMENTALES', 'MATEMÁTICA FUNDAMENTAL',
    'MATEMÁTICAS BÁSICAS', 'LÓGICA Y RAZONAMIENTO', 'CÁLCULO INTEGRAL', 'CÁLCULO DIFERENCIAL', 'CÁLCULO'
]

def inf_encontrar_columna(df, nombres_posibles):
    for nombre in nombres_posibles:
        nombre_norm = nombre.upper().translate(TILDES_MAP)
        for col in df.columns:
            if str(col).strip().upper().translate(TILDES_MAP) == nombre_norm:
                return col
    return None

def inf_normalizar_nota(nota):
    if pd.isna(nota): return nota
    try:
        return float(str(nota).replace(',', '.'))
    except:
        return nota

def inf_leer_excel(file_obj):
    try:
        validar_archivo_xlsx(file_obj)
        excel_file = pd.ExcelFile(file_obj, engine='openpyxl')
        sheet_names = [name.upper() for name in excel_file.sheet_names]
        if 'ORIGINAL' in sheet_names:
            idx = sheet_names.index('ORIGINAL')
            df = pd.read_excel(excel_file, sheet_name=excel_file.sheet_names[idx])
        else:
            df = pd.read_excel(excel_file, sheet_name=0)
        return df
    except Exception as e:
        raise Exception(f"Error al leer el archivo Excel: {str(e)}")

def inf_crear_hoja_general(df_original):
    df_general = pd.DataFrame()
    for col_norm, nombres in MAPEO_COLUMNAS_INF.items():
        col_enc = inf_encontrar_columna(df_original, nombres)
        if col_enc:
            df_general[col_norm] = df_original[col_enc].copy()
    if 'NOTA' in df_general.columns:
        df_general['NOTA'] = df_general['NOTA'].apply(inf_normalizar_nota)
    return df_general

def inf_filtrar_por_materias(df_general, materias):
    if 'MATERIA' not in df_general.columns:
        return pd.DataFrame()
    df_normalizado = df_general['MATERIA'].astype(str).str.upper().str.strip().str.translate(TILDES_MAP)
    materias_normalizado = [str(m).upper().strip().translate(TILDES_MAP) for m in materias]
    mask = df_normalizado.isin(materias_normalizado)
    return df_general[mask].copy()

def inf_aplicar_formato(workbook, nombre_hoja):
    if nombre_hoja not in workbook.sheetnames:
        return
    sheet = workbook[nombre_hoja]
    columnas_nombres = {}
    if sheet.max_row > 0:
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value is not None:
                columnas_nombres[str(cell.value).upper().strip()] = idx
    if sheet.max_row > 0:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    if sheet.max_row > 0:
        sheet.auto_filter.ref = sheet.dimensions
    formato_numero = '0'
    if 'DOCUMENTO' in columnas_nombres and sheet.max_row > 1:
        col_documento = columnas_nombres['DOCUMENTO']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_documento)
            if cell.value is not None:
                try:
                    cell.value = float(str(cell.value).replace(',', '.'))
                    cell.number_format = formato_numero
                except:
                    pass
    if 'GRUPO' in columnas_nombres and sheet.max_row > 1:
        col_grupo = columnas_nombres['GRUPO']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_grupo)
            if cell.value is not None:
                cell.value = str(cell.value).strip()
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left', vertical='center')
    if 'N_PERDIDAS' in columnas_nombres and sheet.max_row > 1:
        col_n_perdidas = columnas_nombres['N_PERDIDAS']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_n_perdidas)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 18

def inf_procesar_archivo(file_obj):
    df_original = inf_leer_excel(file_obj)
    df_general = inf_crear_hoja_general(df_original)
    if df_general.empty:
        raise Exception("No se pudieron mapear las columnas del archivo original.")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_general.to_excel(writer, sheet_name='GENERAL', index=False)
        df_comunicacion = inf_filtrar_por_materias(df_general, MATERIAS_COMUNICACION_INF)
        if not df_comunicacion.empty:
            df_comunicacion.to_excel(writer, sheet_name='COMUNICACION', index=False)
        else:
            df_general.head(0).to_excel(writer, sheet_name='COMUNICACION', index=False)
        df_matematicas = inf_filtrar_por_materias(df_general, MATERIAS_MATEMATICAS_INF)
        if not df_matematicas.empty:
            df_matematicas.to_excel(writer, sheet_name='MATEMATICAS', index=False)
        else:
            df_general.head(0).to_excel(writer, sheet_name='MATEMATICAS', index=False)
    output.seek(0)
    workbook = openpyxl.load_workbook(output)
    inf_aplicar_formato(workbook, 'GENERAL')
    inf_aplicar_formato(workbook, 'COMUNICACION')
    inf_aplicar_formato(workbook, 'MATEMATICAS')
    final_output = io.BytesIO()
    workbook.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()


# ============================================================
# ==================  MÓDULO 5: ESTUDIANTES SUPABASE  ========
# ============================================================

SEMESTRE_MAP_SUP = {'PRIMERO': '1', 'SEGUNDO': '2', 'TERCERO': '3'}
TILDES_MAP_SUP = str.maketrans('ÁÉÍÓÚáéíóúÑñÜü', 'AEIOUaeiouNnUu')

def sup_limpiar_texto(texto):
    if pd.isna(texto): return texto
    texto_limpio = str(texto).translate(TILDES_MAP_SUP)
    texto_limpio = re.sub(r'\s+', ' ', texto_limpio).strip()
    return texto_limpio

def sup_leer_hojas(file_obj):
    try:
        validar_archivo_xlsx(file_obj)
        excel_file = pd.ExcelFile(file_obj, engine='openpyxl')
        sheet_names_upper = [s.upper() for s in excel_file.sheet_names]
        dfs = []
        for nombre_hoja in ['COMUNICACION', 'MATEMATICAS']:
            if nombre_hoja in sheet_names_upper:
                idx = sheet_names_upper.index(nombre_hoja)
                df = pd.read_excel(excel_file, sheet_name=excel_file.sheet_names[idx], dtype=str)
                dfs.append(df)
        if not dfs:
            raise Exception("No se encontraron las hojas COMUNICACION ni MATEMATICAS en el archivo.")
        return pd.concat(dfs, ignore_index=True)
    except Exception as e:
        raise Exception(f"Error al leer el archivo: {str(e)}")

def sup_encontrar_columna(df, nombres_posibles):
    for nombre in nombres_posibles:
        nombre_norm = nombre.upper().translate(TILDES_MAP)
        for col in df.columns:
            if str(col).strip().upper().translate(TILDES_MAP) == nombre_norm:
                return col
    return None

def sup_procesar_archivo(file_obj):
    df = sup_leer_hojas(file_obj)
    col_documento = sup_encontrar_columna(df, ['DOCUMENTO', 'DOCUM', 'PEGE_DOCUMENTOIDENTIDAD', 'DOC'])
    col_nombre = sup_encontrar_columna(df, ['NOMBRE', 'NOM', 'ESTUDIANTE'])
    col_programa = sup_encontrar_columna(df, ['PROGRAMA', 'PROG', 'PROG_NOMBRE'])
    col_sede = sup_encontrar_columna(df, ['SEDE', 'FRANJA'])
    col_facultad = sup_encontrar_columna(df, ['FACULTAD'])
    col_semestre = sup_encontrar_columna(df, ['SEMESTRE', 'SEM'])
    columnas_requeridas = {
        'documento': col_documento, 'nombre': col_nombre, 'facultad': col_facultad,
        'programa': col_programa, 'sede': col_sede, 'semestre': col_semestre
    }
    faltantes = [k for k, v in columnas_requeridas.items() if v is None]
    if faltantes:
        raise Exception(f"No se encontraron las columnas: {', '.join(faltantes)}")
    df_work = pd.DataFrame()
    df_work['_documento'] = df[col_documento]
    df_work['_nombre'] = df[col_nombre]
    df_work['_facultad'] = df[col_facultad] if col_facultad else None
    df_work['_programa'] = df[col_programa]
    df_work['_sede'] = df[col_sede]
    df_work['_semestre'] = df[col_semestre]
    df_work = df_work.drop_duplicates(subset='_documento').copy()
    if df_work.empty:
        raise Exception("No hay registros para procesar.")
    df_work['semestre'] = df_work['_semestre'].apply(
        lambda x: SEMESTRE_MAP_SUP.get(str(x).strip().upper(), str(x).strip()) if pd.notna(x) else x
    )
    df_final = pd.DataFrame()
    df_final['documento'] = df_work['_documento']
    df_final['nombre'] = df_work['_nombre'].apply(sup_limpiar_texto)
    df_final['facultad'] = df_work['_facultad'].values
    df_final['programa'] = df_work['_programa'].values
    df_final['sede'] = df_work['_sede']
    df_final['semestre'] = df_work['semestre']
    df_final = df_final.reset_index(drop=True)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name='ESTUDIANTES', index=False)
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    ws = wb['ESTUDIANTES']
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value is not None:
            try:
                cell.value = float(str(cell.value).replace(',', '.'))
                cell.number_format = '0'
            except:
                pass
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()

# ============================================================
# ==================  MÓDULO 6: VISITAS A GRUPOS  ============
# ============================================================


def vis_aplicar_formato(workbook, nombre_hoja):
    if nombre_hoja not in workbook.sheetnames:
        return
    sheet = workbook[nombre_hoja]
    columnas_nombres = {}
    if sheet.max_row > 0:
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value is not None:
                columnas_nombres[str(cell.value).upper().strip()] = idx
    if sheet.max_row > 0:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    if sheet.max_row > 0:
        sheet.auto_filter.ref = sheet.dimensions
    if 'GRUPO' in columnas_nombres and sheet.max_row > 1:
        col_grupo = columnas_nombres['GRUPO']
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_grupo)
            if cell.value is not None:
                cell.value = str(cell.value).strip()
                cell.number_format = '@'
                cell.alignment = Alignment(horizontal='left', vertical='center')
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 18

def vis_detectar_dia(nombre_hoja, dias_validos):
    """
    Toma la primera palabra del nombre de la hoja, le quita tildes/mayúsculas,
    y verifica si corresponde a un día válido. Ej: 'Lunes Noche' -> 'LUNES'.
    Retorna el día normalizado o None si no coincide.
    """
    nombre_norm = str(nombre_hoja).strip().upper().translate(TILDES_MAP)
    if not nombre_norm:
        return None
    primer_token = nombre_norm.split()[0]
    if primer_token in dias_validos:
        return primer_token
    return None

def vis_determinar_jornada(horario):
    """
    Analiza la hora de inicio (primeros 4 dígitos) del valor de HORARIO
    (ej: '0800 - 1200') y determina la JORNADA correspondiente:
    0700-1159 -> MAÑANA | 1200-1759 -> TARDE | 1800-2159 -> NOCHE
    No modifica el valor original de HORARIO.
    """
    if pd.isna(horario):
        return None
    texto = str(horario).strip()
    match = re.match(r'^(\d{3,4})', texto)
    if not match:
        return None
    hora_inicio = match.group(1).zfill(4)
    try:
        hora_num = int(hora_inicio)
    except:
        return None
    if 700 <= hora_num <= 1159:
        return 'MAÑANA'
    elif 1200 <= hora_num <= 1759:
        return 'TARDE'
    elif 1800 <= hora_num <= 2159:
        return 'NOCHE'
    else:
        return None

def vis_procesar_archivo(archivos):
    """
    archivos: lista de file objects (1, 2 o 3 archivos).
    """
    DIAS_VALIDOS = {'LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO'}
    COLUMNAS_GENERAL = ['DIA', 'HORARIO', 'SALON', 'GRUPO', 'JORNADA', 'ASIGNATURA', 'DOCENTE']
    COLUMNAS_VISITAS = ['DIA', 'HORARIO', 'SALON', 'GRUPO', 'SEMESTRE', 'JORNADA']
    COLUMNAS_ORIGINALES = ['SALON', 'GRUPO', 'ASIGNATURA', 'DOCENTE', 'HORARIO']
    FILAS_HEADER = [6, 5, 7, 0]  # fila 7, fila 6, fila 8, fila 1 (índice 0-based)

    generales = []

    for file_obj in archivos:
        validar_archivo_xlsx(file_obj)
        excel_file = pd.ExcelFile(file_obj, engine='openpyxl')

        for nombre_hoja in excel_file.sheet_names:
            dia_detectado = vis_detectar_dia(nombre_hoja, DIAS_VALIDOS)
            if dia_detectado is None:
                continue

            df_hoja = None
            for header_idx in FILAS_HEADER:
                df_temp = pd.read_excel(excel_file, sheet_name=nombre_hoja, header=header_idx, dtype=str)
                df_temp.columns = [str(c).strip().upper().translate(str.maketrans('ÁÉÍÓÚÑÜ', 'AEIOUNU')) for c in df_temp.columns]
                columnas_encontradas = [c for c in COLUMNAS_ORIGINALES if c in df_temp.columns]
                if len(columnas_encontradas) >= 2:
                    df_hoja = df_temp
                    break

            if df_hoja is None:
                continue

            df_hoja['DIA'] = dia_detectado
            if 'HORARIO' in df_hoja.columns:
                df_hoja['JORNADA'] = df_hoja['HORARIO'].apply(vis_determinar_jornada)
            else:
                df_hoja['JORNADA'] = None

            for col_critica in ['SALON', 'GRUPO', 'HORARIO']:
                if col_critica in df_hoja.columns:
                    df_hoja = df_hoja[
                        df_hoja[col_critica].notna() &
                        (df_hoja[col_critica].astype(str).str.strip() != '') &
                        (df_hoja[col_critica].astype(str).str.strip().str.upper() != 'NAN')
                    ]

            generales.append(df_hoja)

    if not generales:
        raise Exception("No se encontraron hojas válidas (LUNES a SABADO) en ninguno de los archivos.")

    df_general_raw = pd.concat(generales, ignore_index=True)

    df_general = pd.DataFrame()
    for col in COLUMNAS_GENERAL:
        if col in df_general_raw.columns:
            df_general[col] = df_general_raw[col]
        else:
            df_general[col] = None

    df_general = df_general.reset_index(drop=True)

    columnas_duplicado = ['DIA', 'HORARIO', 'GRUPO']
    if all(c in df_general.columns for c in columnas_duplicado):
        df_general = df_general.drop_duplicates(subset=columnas_duplicado, keep='first')
        df_general = df_general.reset_index(drop=True)

    df_visitas = df_general.copy()
    df_visitas['SEMESTRE'] = df_visitas['GRUPO'].apply(determinar_semestre)
    df_visitas = df_visitas[df_visitas['SEMESTRE'].notna()].copy()
    df_visitas = df_visitas.reset_index(drop=True)
    df_visitas = df_visitas[[c for c in COLUMNAS_VISITAS if c in df_visitas.columns]]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_general.to_excel(writer, sheet_name='GENERAL', index=False)
        df_visitas.to_excel(writer, sheet_name='VISITAS', index=False)

    output.seek(0)
    workbook = openpyxl.load_workbook(output)
    vis_aplicar_formato(workbook, 'GENERAL')
    vis_aplicar_formato(workbook, 'VISITAS')

    final_output = io.BytesIO()
    workbook.save(final_output)
    final_output.seek(0)
    return final_output.getvalue()
    
# ============================================================
# ==================  INTERFAZ STREAMLIT  ====================
# ============================================================

# Sidebar
with st.sidebar:
    st.markdown('<div class="sidebar-title">🎓 Procesador Académico</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-subtitle">Sistema de gestión de datos</div>', unsafe_allow_html=True)
    st.markdown("---")

    modulo = st.radio(
        "Selecciona un módulo:",
        options=[
            "Facultades",
            "Matriculados",
            "Cruce",
            "Informe General",
            "Estudiantes Supabase",
            "Visitas a Grupos"
        ],
        label_visibility="collapsed"
    )

    st.markdown("---")
    st.markdown(
        '<div style="font-size:0.72rem; color:rgba(255,255,255,0.25); line-height:1.6;">'
        'Los archivos se procesan en memoria.<br>Ningún dato se almacena en el servidor.'
        '</div>',
        unsafe_allow_html=True
    )


# ─────────────────────────────────────────────
#  MÓDULO 1 — FACULTADES
# ─────────────────────────────────────────────
if modulo == "Facultades":
    st.markdown("""
    <div class="module-header">
        <h1>Procesador de Facultades</h1>
        <p>Lee el archivo de facultades y genera un libro Excel por cada facultad,<br>
        con hoja GENERAL y una hoja por programa. También produce el <strong>INFORME GENERAL</strong>.</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-card">
        <strong>¿Qué hace este módulo?</strong><br>
        • Filtra estudiantes con nota entre <strong>0.0 y 2.9</strong> (reprobados).<br>
        • Genera un archivo Excel por cada facultad encontrada.<br>
        • Cada archivo contiene hoja GENERAL + una hoja por programa académico.<br>
        • Genera adicionalmente el archivo <strong>INFORME GENERAL.xlsx</strong> con todas las facultades unidas.<br>
        • Corrige ortografía automáticamente (tildes en nombres de materias y programas).
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="warning-card">
        <strong>Columnas Necesarias</strong><br>
        <code>DOCUMENTO / DOC / DOCUM / PEGE_DOCUMENTOIDENTIDAD</code>
        <code>NOMBRE / NOM / ESTUDIANTE</code>
        <code>FACULTAD</code>
        <code>PROGRAMA / PROG / PROG_NOMBRE</code>
        <code>MATERIA</code>
        <code>EVALUACION</code>
        <code>GRUPO</code>
        <code>NOTA</code>
        <code>SEM / SEMESTRE</code><br><br>
        La columna <strong>N_PERDIDAS</strong> se genera automáticamente.
    </div>
    """, unsafe_allow_html=True)

    col_fac1, col_fac2 = st.columns(2)
    with col_fac1:
        st.markdown('<div class="upload-label">📂 Archivo 1 (obligatorio)</div>', unsafe_allow_html=True)
        archivo_fac = st.file_uploader(
            "Archivo Facultades 1",
            type=["xlsx"],
            key="uploader_fac",
            label_visibility="collapsed"
        )
    with col_fac2:
        st.markdown('<div class="upload-label">📂 Archivo 2 (opcional)</div>', unsafe_allow_html=True)
        archivo_fac2 = st.file_uploader(
            "Archivo Facultades 2",
            type=["xlsx"],
            key="uploader_fac2",
            label_visibility="collapsed"
        )

    st.markdown("---")
    st.markdown("**⚙️ Opciones de filtrado de grupos**")

    col1_fac, col2_fac, col3_fac = st.columns(3)
    with col1_fac:
        calendario_a_fac = st.checkbox(
            "Calendario A",
            help=HELP_CALENDARIO_A,
            key="cal_a_fac"
        )
        if calendario_a_fac:
            st.markdown(f'<span class="badge badge-blue">{BADGE_CALENDARIO_A}</span>', unsafe_allow_html=True)
    with col2_fac:
        calendario_b_fac = st.checkbox(
            "Calendario B",
            help=HELP_CALENDARIO_B,
            key="cal_b_fac"
        )
        if calendario_b_fac:
            st.markdown(f'<span class="badge badge-blue">{BADGE_CALENDARIO_B}</span>', unsafe_allow_html=True)
    with col3_fac:
        calendario_c_fac = st.checkbox(
            "Calendario C",
            help=HELP_CALENDARIO_C,
            key="cal_c_fac"
        )
        if calendario_c_fac:
            st.markdown(f'<span class="badge badge-blue">{BADGE_CALENDARIO_C}</span>', unsafe_allow_html=True)

    st.markdown("---")

    if archivo_fac:
        if st.button("▶  Procesar Facultades", key="btn_fac", use_container_width=True):
            with st.spinner("Procesando facultades..."):
                try:
                    archivos = fac_procesar_archivo(archivo_fac, archivo_fac2, calendario_a_fac, calendario_b_fac, calendario_c_fac)
                    st.success(f"✅ Proceso completado. Se generaron **{len(archivos)} archivos**.")

                    import zipfile
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for nombre_archivo, contenido in archivos.items():
                            zip_file.writestr(nombre_archivo, contenido)
                    zip_buffer.seek(0)

                    st.download_button(
                        label="⬇️  Descargar todos los archivos (.zip)",
                        data=zip_buffer.getvalue(),
                        file_name="FACULTADES_PROCESADAS.zip",
                        mime="application/zip",
                        key="dl_fac_zip"
                    )
                except Exception as e:
                    st.error(f"❌ Error al procesar: {str(e)}")
    else:
        st.markdown("""
        <div class="warning-card">
            ⬆️ Sube el archivo Excel para habilitar el procesamiento.
        </div>
        """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  MÓDULO 2 — MATRICULADOS
# ─────────────────────────────────────────────
elif modulo == "Matriculados":
    st.markdown("""
    <div class="module-header">
        <h1>Procesador de Matriculados</h1>
        <p>Procesa el archivo de matriculados y genera hojas organizadas:<br>
        <strong>GENERAL · MATRICULADOS · COMUNICACION · MATEMATICAS</strong></p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-card">
        <strong>¿Qué hace este módulo?</strong><br>
        • Filtra y organiza los datos de estudiantes matriculados.<br>
        • Asigna <strong>FACULTAD</strong> automáticamente según el programa académico.<br>
        • Normaliza la columna <strong>SEDE</strong> (NORTE / SUR).<br>
        • Filtra grupos válidos y determina el semestre (PRIMERO / SEGUNDO / TERCERO).<br>
        • Genera hoja <strong>MATRICULADOS</strong> con conteo de estudiantes únicos por grupo.<br>
        • Genera hojas separadas de <strong>COMUNICACION</strong> y <strong>MATEMATICAS</strong>.
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="warning-card">
        <strong>Columnas Necesarias</strong><br>
        <code>DOCUMENTO / DOC / DOCUM / PEGE_DOCUMENTOIDENTIDAD</code>
        <code>NOMBRE / NOM / ESTUDIANTE</code>
        <code>PROGRAMA / PROG / PROG_NOMBRE</code>
        <code>SEDE / FRANJA</code>
        <code>MATERIA</code>
        <code>GRUPO</code>
        <code>SEMESTRE</code>
        <code>CELULAR</code>
        <code>CORREO_INST</code>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="upload-label">📂 Archivo de entrada</div>', unsafe_allow_html=True)
    archivo_mat = st.file_uploader(
        "Sube el archivo Excel de Matriculados",
        type=["xlsx"],
        key="uploader_mat",
        label_visibility="collapsed"
    )

    st.markdown("---")

    if archivo_mat:
        if st.button("▶  Procesar Matriculados", key="btn_mat", use_container_width=True):
            with st.spinner("Procesando matriculados..."):
                try:
                    resultado = mat_procesar_archivo(archivo_mat)
                    st.success("✅ Proceso completado exitosamente.")
                    st.download_button(
                        label="⬇️  Descargar archivo procesado",
                        data=resultado,
                        file_name="MATRICULADOS_PROCESADO.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_mat"
                    )
                except Exception as e:
                    st.error(f"❌ Error al procesar: {str(e)}")
    else:
        st.markdown("""
        <div class="warning-card">
            ⬆️ Sube el archivo Excel para habilitar el procesamiento.
        </div>
        """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  MÓDULO 3 — CRUCE
# ─────────────────────────────────────────────
elif modulo == "Cruce":
    st.markdown("""
    <div class="module-header">
        <h1>Procesador de Cruce</h1>
        <p>Cruza los archivos <strong>TUTORIAS</strong>, <strong>MATRICULADOS</strong> e <strong>INFORME GENERAL</strong><br>
        para generar hojas de asistencia y estudiantes que ganaron.</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-card">
        <strong>¿Qué hace este módulo?</strong><br>
        • Filtra por AREA: Matemáticas (M, DCB) y Comunicación (C).<br>
        • Cruza con MATRICULADOS para obtener GRUPO y SEDE por estudiante.<br>
        • Genera hojas: <br>
        <strong>GENERAL · ASISTENCIA MATEMÁTICAS · ASISTENCIA COMUNICACIÓN · ASISTENCIA PMA · GANARON MATEMÁTICAS · GANARON COMUNICACIÓN</strong>.<br>
        • GANARON = Estudiantes que asistieron pero no aparecen en el INFORME GENERAL (No reprobaron).
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="warning-card">
        <strong>Archivos Necesarios</strong><br>
        TUTORIAS — descargado desde el Panel Admin del formulario ("Descargar Informe").<br>
        MATRICULADOS — archivo ya procesado por el módulo "Matriculados".<br>
        INFORME GENERAL — archivo ya procesado por el módulo "Informe General".
    </div>
    """, unsafe_allow_html=True)

    col_c1, col_c2, col_c3 = st.columns(3)
    with col_c1:
        st.markdown('<div class="upload-label">📂 Archivo TUTORIAS</div>', unsafe_allow_html=True)
        archivo_cruce = st.file_uploader("Archivo CRUCE", type=["xlsx"], key="uploader_cruce", label_visibility="collapsed")
    with col_c2:
        st.markdown('<div class="upload-label">📂 Archivo MATRICULADOS</div>', unsafe_allow_html=True)
        archivo_mat_cruce = st.file_uploader("Archivo MATRICULADOS", type=["xlsx"], key="uploader_mat_cruce", label_visibility="collapsed")
    with col_c3:
        st.markdown('<div class="upload-label">📂 Archivo INFORME GENERAL</div>', unsafe_allow_html=True)
        archivo_informe_cruce = st.file_uploader("Archivo INFORME GENERAL", type=["xlsx"], key="uploader_inf_cruce", label_visibility="collapsed")

    st.markdown("---")
    st.markdown("**⚙️ Opciones de filtrado de grupos**")

    col1_c, col2_c, col3_c = st.columns(3)
    with col1_c:
        calendario_a_cruce = st.checkbox(
            "Calendario A",
            help=HELP_CALENDARIO_A,
            key="cal_a_cruce"
        )
        if calendario_a_cruce:
            st.markdown(f'<span class="badge badge-blue">{BADGE_CALENDARIO_A}</span>', unsafe_allow_html=True)
    with col2_c:
        calendario_b_cruce = st.checkbox(
            "Calendario B",
            help=HELP_CALENDARIO_B,
            key="cal_b_cruce"
        )
        if calendario_b_cruce:
            st.markdown(f'<span class="badge badge-blue">{BADGE_CALENDARIO_B}</span>', unsafe_allow_html=True)
    with col3_c:
        calendario_c_cruce = st.checkbox(
            "Calendario C",
            help=HELP_CALENDARIO_C,
            key="cal_c_cruce"
        )
        if calendario_c_cruce:
            st.markdown(f'<span class="badge badge-blue">{BADGE_CALENDARIO_C}</span>', unsafe_allow_html=True)
    st.markdown("---")

    archivos_listos = archivo_cruce and archivo_mat_cruce and archivo_informe_cruce

    if archivos_listos:
        if st.button("▶  Procesar Cruce", key="btn_cruce", use_container_width=True):
            with st.spinner("Procesando cruce de archivos..."):
                try:
                    resultado = cruce_procesar_archivo(
                        archivo_cruce, archivo_mat_cruce, archivo_informe_cruce,
                        calendario_a_cruce, calendario_b_cruce, calendario_c_cruce
                    )
                    st.success("✅ Proceso completado exitosamente.")
                    st.download_button(
                        label="⬇️  Descargar archivo procesado",
                        data=resultado,
                        file_name="CRUCE_PROCESADO.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_cruce"
                    )
                except Exception as e:
                    st.error(f"❌ Error al procesar: {str(e)}")
    else:
        faltantes_c = []
        if not archivo_cruce: faltantes_c.append("CRUCE")
        if not archivo_mat_cruce: faltantes_c.append("MATRICULADOS")
        if not archivo_informe_cruce: faltantes_c.append("INFORME GENERAL")
        st.markdown(f"""
        <div class="warning-card">
            ⬆️ Faltan archivos por subir: <strong>{' · '.join(faltantes_c)}</strong>
        </div>
        """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  MÓDULO 4 — INFORME GENERAL
# ─────────────────────────────────────────────
elif modulo == "Informe General":
    st.markdown("""
    <div class="module-header">
        <h1>Procesador de Informe General</h1>
        <p>Procesa el archivo de informe general y genera hojas organizadas:<br>
        <strong>GENERAL · COMUNICACION · MATEMATICAS</strong></p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-card">
        <strong>¿Qué hace este módulo?</strong><br>
        • Lee la hoja <strong>ORIGINAL</strong> del archivo (o la primera hoja si no existe).<br>
        • Normaliza y mapea las columnas al formato estándar.<br>
        • Genera hoja <strong>GENERAL</strong> con todas las materias.<br>
        • Filtra y genera hojas separadas de <strong>COMUNICACION</strong> y <strong>MATEMATICAS</strong>.<br>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="warning-card">
        <strong>Archivo Necesario</strong><br>
        Debe subirse el Archivo "Informe General" generado en el modulo de "Procesador de Facultades".
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="upload-label">📂 Archivo de entrada (INFORME GENERAL CREADO POR "PROCESADOR DE FACULTADES")</div>', unsafe_allow_html=True)
    archivo_inf = st.file_uploader(
        "Sube el archivo Excel de Informe General",
        type=["xlsx"],
        key="uploader_inf",
        label_visibility="collapsed"
    )

    st.markdown("---")

    if archivo_inf:
        if st.button("▶  Procesar Informe General", key="btn_inf", use_container_width=True):
            with st.spinner("Procesando informe general..."):
                try:
                    resultado = inf_procesar_archivo(archivo_inf)
                    st.success("✅ Proceso completado exitosamente.")
                    st.download_button(
                        label="⬇️  Descargar archivo procesado",
                        data=resultado,
                        file_name="INFORME_GENERAL_PROCESADO.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_inf"
                    )
                except Exception as e:
                    st.error(f"❌ Error al procesar: {str(e)}")
    else:
        st.markdown("""
        <div class="warning-card">
            ⬆️ Sube el archivo Excel para habilitar el procesamiento.
        </div>
        """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  MÓDULO 5 — ESTUDIANTES SUPABASE
# ─────────────────────────────────────────────
elif modulo == "Estudiantes Supabase":
    st.markdown("""
    <div class="module-header">
        <h1>Estudiantes para Supabase</h1>
        <p>Prepara los datos de estudiantes para importación a Supabase:<br></p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-card">
        <strong>¿Qué hace este módulo?</strong><br>
        • Lee las hojas <strong>COMUNICACION</strong> y <strong>MATEMATICAS</strong> del archivo.<br>
        • Elimina duplicados por número de documento.<br>
        • Convierte el semestre en número.
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="warning-card">
        <strong>Archivo Necesario</strong><br>
        Debe subirse el archivo MATRICULADOS PROCESADO generado en el módulo de "Procesador de Matriculados".
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="upload-label">📂 Archivo de entrada (MATRICULADOS PROCESADO)</div>', unsafe_allow_html=True)
    archivo_sup = st.file_uploader(
        "Sube el archivo Excel con hojas COMUNICACION y MATEMATICAS",
        type=["xlsx"],
        key="uploader_sup",
        label_visibility="collapsed"
    )

    st.markdown("---")

    if archivo_sup:
        if st.button("▶  Procesar Estudiantes Supabase", key="btn_sup", use_container_width=True):
            with st.spinner("Preparando datos para Supabase..."):
                try:
                    resultado = sup_procesar_archivo(archivo_sup)
                    st.success("✅ Proceso completado exitosamente.")
                    st.download_button(
                        label="⬇️  Descargar archivo ESTUDIANTES",
                        data=resultado,
                        file_name="ESTUDIANTES_SUPABASE.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_sup"
                    )
                except Exception as e:
                    st.error(f"❌ Error al procesar: {str(e)}")
    else:
        st.markdown("""
        <div class="warning-card">
            ⬆️ Sube el archivo Excel para habilitar el procesamiento.
        </div>
        """, unsafe_allow_html=True)



# ─────────────────────────────────────────────
#  MÓDULO 6 — VISITAS A GRUPOS
# ─────────────────────────────────────────────

elif modulo == "Visitas a Grupos":
    st.markdown("""
    <div class="module-header">
        <h1>Visitas a Grupos</h1>
        <p>Une todas las hojas válidas de los archivos subidos en una hoja <strong>GENERAL</strong>,<br>
        y genera la hoja <strong>VISITAS</strong> con semestre asignado según el grupo.</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-card">
        <strong>¿Qué hace este módulo?</strong><br>
        • Lee únicamente las hojas llamadas <strong>LUNES, MARTES, MIERCOLES, JUEVES, VIERNES o SABADO</strong>.<br>
        • Los encabezados se buscan automáticamente en las filas 7, 6, 8 y 1.<br>
        • Los encabezados del archivo original pueden estar en minúscula o mayúscula — se normalizan automáticamente.<br>
        • Une los archivos en una sola <strong>GENERAL</strong>, luego genera <strong>VISITAS</strong> con semestre asignado.<br>
        • Puedes subir entre <strong>1 y 3 archivos</strong>. No es obligatorio subir los 3.
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="warning-card">
        <strong>Columnas Necesarias</strong><br>
        <code>SALON</code>
        <code>GRUPO</code>
        <code>ASIGNATURA</code>
        <code>DOCENTE</code>
        <code>HORARIO</code><br><br>
        Las columnas <strong>DIA</strong>, <strong>JORNADA</strong> y <strong>SEMESTRE</strong> se generan automáticamente.
    </div>
    """, unsafe_allow_html=True)

    col_v1, col_v2, col_v3 = st.columns(3)
    with col_v1:
        st.markdown('<div class="upload-label">📂 1 — DIURNA</div>', unsafe_allow_html=True)
        archivo_vis_1 = st.file_uploader("Archivo 1", type=["xlsx"], key="uploader_vis_1", label_visibility="collapsed")
    with col_v2:
        st.markdown('<div class="upload-label">📂 2 — DIURNA (opcional)</div>', unsafe_allow_html=True)
        archivo_vis_2 = st.file_uploader("Archivo 2", type=["xlsx"], key="uploader_vis_2", label_visibility="collapsed")
    with col_v3:
        st.markdown('<div class="upload-label">📂 3 — NOCTURNA (opcional)</div>', unsafe_allow_html=True)
        archivo_vis_3 = st.file_uploader("Archivo 3", type=["xlsx"], key="uploader_vis_3", label_visibility="collapsed")

    st.markdown("---")

    archivos_vis = [f for f in [archivo_vis_1, archivo_vis_2, archivo_vis_3] if f is not None]

    if archivos_vis:
        if st.button("▶  Procesar Visitas a Grupos", key="btn_vis", use_container_width=True):
            with st.spinner("Procesando visitas a grupos..."):
                try:
                    resultado = vis_procesar_archivo(archivos_vis)
                    st.success(f"✅ Proceso completado. Se procesaron **{len(archivos_vis)} archivo(s)**.")
                    st.download_button(
                        label="⬇️  Descargar archivo procesado",
                        data=resultado,
                        file_name="VISITAS_GRUPOS_PROCESADO.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_vis"
                    )
                except Exception as e:
                    st.error(f"❌ Error al procesar: {str(e)}")
    else:
        st.markdown("""
        <div class="warning-card">
            ⬆️ Sube al menos un archivo Excel para habilitar el procesamiento.
        </div>
        """, unsafe_allow_html=True)
